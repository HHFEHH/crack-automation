import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, PatternFill
from PIL import Image as IMG
from .models import Category,Crack,CrackObj

def facility(wb,pk):
  baseWidth = 500
  baseHeight = 400
  category = Category.objects.get(pk=pk)
  
  grayFill = PatternFill(start_color='CCCCCC',
                        end_color='CCCCCC', fill_type='solid')
  sheet = wb.worksheets[0]
  sheet.title = '시설물 현황'

  sheet = wb['시설물 현황']
  sheet['B2'] = "□ 시설물 현황"
  sheet['B3'] = "가. 일반현황"
  sheet['B4'] = '시설물명'
  sheet['B5'] = '시설물위치'
  sheet['B6'] = '용도'
  sheet['B7'] = '구조형식'
  sheet['B8'] = '종별'
  sheet['B9'] = '규모 및 제원 추가사항'
  sheet['B12'] = "나. 전경사진"

  sheet['E4'] = "시설물번호"
  sheet['E5'] = "준공일자"
  sheet['E6'] = "시설물규모"
  sheet['E7'] = "부대시설"

  sheet['D8'] = '전차안전등급'
  sheet['F8'] = '점검결과안전등급'

  sheet['B4'].fill = grayFill
  sheet['B5'].fill = grayFill
  sheet['B6'].fill = grayFill
  sheet['B7'].fill = grayFill
  sheet['B9'].fill = grayFill

  sheet['E4'].fill = grayFill
  sheet['E5'].fill = grayFill
  sheet['E6'].fill = grayFill
  sheet['E7'].fill = grayFill

  sheet['B8'].fill = grayFill
  sheet['D8'].fill = grayFill
  sheet['F8'].fill = grayFill

  sheet.merge_cells('C4:D4')
  sheet.merge_cells('C5:D5')
  sheet.merge_cells('C6:D6')
  sheet.merge_cells('C7:D7')

  sheet.merge_cells('F4:G4')
  sheet.merge_cells('F5:G5')
  sheet.merge_cells('F6:G6')
  sheet.merge_cells('F7:G7')

  
  sheet.merge_cells('F4:G4')
  sheet.merge_cells('F5:G5')
  sheet.merge_cells('F6:G6')
  sheet.merge_cells('F7:G7')

  sheet.merge_cells('B9:G9')
  sheet.merge_cells('B10:G10')

  sheet['C4'] = category.facilityName
  sheet['C5'] = category.facilityNo
  sheet['C6'] = category.usage
  sheet['C7'] = category.structuralForm

  sheet['F4'] = category.facilityNo
  sheet['F5'] = category.completionDate
  sheet['F6'] = category.facilityStructure
  sheet['F7'] = category.amenities


  sheet['C8'] = category.floors
  sheet['E8'] = category.grade
  sheet['G8'] = category.testResults

  sheet['B10'] = category.plus

  sheet['B12'] = '나. 전경사진'
  sheet['B32'] = '다. 위치도'

  frontViewPath = category.frontView.url[1:]
  locationMapPath = category.locationMap.url[1:]

  frontView = IMG.open(frontViewPath)
  frontWidth,frontHeight = frontView.size

  if (frontWidth < frontHeight):
    frontNewWidth = int((frontWidth/frontHeight) * baseHeight)
    frontViewImage = openpyxl.drawing.image.Image(frontViewPath)
    frontViewImage.width = frontNewWidth
    frontViewImage.height = baseHeight
    sheet.add_image(frontViewImage,"B13")
  else:
    frontNewHeight = int((frontHeight/frontWidth) * baseWidth)
    frontViewImage = openpyxl.drawing.image.Image(frontViewPath)
    frontViewImage.width = baseWidth
    frontViewImage.height = frontNewHeight
    sheet.add_image(frontViewImage,"B13")

  
  locationMap = IMG.open(locationMapPath)
  locationWidth,locationHeight = locationMap.size

  if (locationWidth < locationHeight):
    locationNewWidth = int((frontWidth/frontHeight) * baseHeight)
    locationMapImage = openpyxl.drawing.image.Image(frontViewPath)
    locationMapImage.width = locationNewWidth
    locationMapImage.height = baseHeight
    sheet.add_image(locationMapImage,"B33")
  else:
    locationNewHeight = int((locationHeight/locationWidth) * baseWidth)
    locationMapImage = openpyxl.drawing.image.Image(locationMapPath)
    locationMapImage.width = baseWidth
    locationMapImage.height = locationNewHeight
    sheet.add_image(locationMapImage,"B33")

  sheet.sheet_view.view = "pageBreakPreview"
  for row in sheet.rows:
      for cell in row:
          cell.alignment = Alignment(horizontal="center", vertical="center")
  return wb


def looks(wb,pk):
  baseWidth = 210
  imgCell = 2
  infoCell = 10
  cellB = chr(66)
  cellC = chr(67)
  sheet = wb.create_sheet("외관조사사진", 1)
  sheet = wb['외관조사사진']
  category = Category.objects.get(pk=pk)
  cracks = Crack.objects.filter(category__facilityName__icontains=category.facilityName)
  sheet.column_dimensions["A"].width = 1
  sheet.column_dimensions["D"].width = 1
  sheet.column_dimensions["B"].width = 27
  sheet.column_dimensions["C"].width = 27
  sheet.column_dimensions["E"].width = 27
  sheet.column_dimensions["F"].width = 27

  for crack in cracks:
    crackObj = CrackObj.objects.filter(parent=crack.id)
    numbering = crackObj.count()
    if numbering < 3:
      numbering = 0
    else:
      numbering = numbering-2
    crackObj = crackObj[numbering:]
    if crackObj.count() == 2:
      for crackObj in crackObj:
        path = crackObj.image.url[1:]
        img = IMG.open(path) # 사진의 비율을 알기 위한 변수 PIL 라이브러리
        wpercent = baseWidth/float(img.size[0])
        hsize = int((float(img.size[1])* float(wpercent)))

        flatPath = crackObj.flatting_image.url[1:]
        flatImg = IMG.open(flatPath) # 사진의 비율을 알기 위한 변수 PIL 라이브러리
        flatwPercent = baseWidth/float(img.size[0])
        flathSize = int((float(flatImg.size[1])* float(flatwPercent)))

        image = openpyxl.drawing.image.Image(path) # 엑셀에 이미지 삽입을 위한 변수 openpyxl 라이브러리
        flatImage = openpyxl.drawing.image.Image(flatPath)
        
        image.width = baseWidth
        image.height = 130
        
        flatImage.width = baseWidth
        flatImage.height = 130
  
        sheet.add_image(image,cellB + str(imgCell))
        sheet.add_image(flatImage, cellC + str(imgCell))

        sheet[cellB+str(infoCell)] = '사진번호: ' + str(crackObj.id)
        sheet[cellB+str(infoCell+1)] = '위치: ' + str(crack.floor) + str(crack.location)
        sheet[cellB+str(infoCell+2)] = '점검내용: ' + str(crack.desc)
        sheet[cellC+str(infoCell+2)] = '손상규모: ' + str(crackObj.crackLength)
        sheet[cellB+str(infoCell+3)] = '발생원인: ' + str(crack.cause)
        sheet[cellC+str(infoCell+3)] = '진행유무: ' + str(crack.progress)

        cellB = ord(cellB) + 3
        cellB = chr(cellB)
        
        cellC = ord(cellC) + 3
        cellC = chr(cellC)

        if ord(cellB) > 70:
          cellB = chr(66)
          infoCell += 7
          imgCell +=7
        if ord(cellC) > 70:
          cellC = chr(67)
          imgCell +=7
          infoCell += 7
        sheet.sheet_view.view = "pageBreakPreview"
    else:
      for crackObj in crackObj:
        path = crackObj.image.url[1:]
        img = IMG.open(path) # 사진의 비율을 알기 위한 변수 PIL 라이브러리
        wpercent = baseWidth/float(img.size[0])
        hsize = int((float(img.size[1])* float(wpercent)))

        flatPath = crackObj.flatting_image.url[1:]
        flatImg = IMG.open(flatPath) # 사진의 비율을 알기 위한 변수 PIL 라이브러리
        flatwPercent = baseWidth/float(img.size[0])
        flathSize = int((float(flatImg.size[1])* float(flatwPercent)))

        if hsize > 160:
          hsize = 160
          baseWidth = baseWidth * wpercent
        if flathSize > 160:
          flathSize = 160

        image = openpyxl.drawing.image.Image(path) # 엑셀에 이미지 삽입을 위한 변수 openpyxl 라이브러리
        flatImage = openpyxl.drawing.image.Image(flatPath)
          
        image.width = baseWidth
        image.height = hsize
          
        flatImage.width = baseWidth
        flatImage.height = flathSize
    
        sheet.add_image(image,cellB + str(imgCell))
        sheet.add_image(flatImage, cellC + str(imgCell))

        sheet[cellB+str(infoCell)] = '사진번호: ' + str(crackObj.id)
        sheet[cellB+str(infoCell+1)] = '위치: ' + str(crack.floor) + str(crack.location)
        sheet[cellB+str(infoCell+2)] = '점검내용: ' + str(crack.desc)
        sheet[cellC+str(infoCell+2)] = '손상규모: ' + str(crackObj.crackLength)
        sheet[cellB+str(infoCell+3)] = '발생원인: ' + str(crack.cause)
        sheet[cellC+str(infoCell+3)] = '진행유무: ' + str(crack.progress)
        sheet.sheet_view.view = "pageBreakPreview"
  return wb
